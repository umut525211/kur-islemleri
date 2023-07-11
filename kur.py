from openpyxl import Workbook,load_workbook
wb = load_workbook("try.xlsx")
ws = wb.active
tarih=[]
high=[]
low=[]
kur=[]
baslik=[]
#İlk başta tablonun başlık kısmını aldım
for a in ws[1]:
    baslik.append(a.value)
#Bu kısımda tabloda başlık kısmından yararlanarak istediklerimin verilerini aldım ve bu verileri farklı dizilere kaydettim
for sira in ws.iter_rows(min_row=2, values_only=True):
    tarih.append(sira[baslik.index('Date')])
    high.append(sira[baslik.index('High')])
    low.append(sira[baslik.index('Low')])
#Üstte kaydettiğim verilerin bulunduğu dizileri tek diziye kaydettim
for date1, high1, low1 in zip(tarih, high, low):
    kur.append([date1,low1,high1])
    
wb.close()
para=15000
fark=0
maxgun=mingun=sayac=aylik_max=fark_tarih=fark_min=0
aylik_min=100
#Aşağıdaki while döngüsünde her gün farkı için en fazla hangi gün bekleyerek kâr ve zarar edebileceğini hesaplatıp tek tek kaydettim
while fark<len(kur):
    sayac=0
    while sayac<(len(kur)-fark):
        deger=kur[sayac+fark][2]*(para/kur[sayac][1])-para
        deger2=kur[sayac+fark][1]*(para/kur[sayac][2])-para
        if aylik_max<deger:
            aylik_max=deger
            fark_tarih=sayac
            maxgun=fark
        if aylik_min>deger2:
            aylik_min=deger2
            fark_min=sayac
            mingun=fark
        sayac=sayac+1
    fark=fark+1
#Ekrana ne kadar gün beklerse en çok kâr elde edeceğini yazdırdım ve ne kadar beklerse yapacağı en büyük zararı ekrana yazdırdım 
print(kur[fark_tarih][0],"/",kur[fark_tarih+maxgun][0],"Tarihleri Arasında",maxgun,"İş Günü Kadar Bekleyerek Yapabilecek En Yüksek Kâr: ",aylik_max)
print(kur[fark_min][0],"/",kur[fark_min+mingun][0],"Tarihleri Arasında",mingun,"İş Günü Kadar Bekleyerek Yapabilecek En Düşük Kâr: ",aylik_min)

#Girilen fark  değerine göre hesaplama
print()
fark=int(input("Farkı girin(Max=261, Min=0): "))
print()
sayac=aylik_max=fark_tarih=fark_min=0
aylik_min=10000
#Girilen fark sayısına göre yapabileceği max ve min kâr miktarını hesaplatıp ekrana yazdırdım.
while sayac<(len(kur)-fark):
    deger=kur[sayac+fark][2]*(para/kur[sayac][1])-para
    deger2=kur[sayac+fark][1]*(para/kur[sayac][2])-para
    if aylik_max<deger:
        aylik_max=deger
        fark_tarih=sayac
    if aylik_min>deger2:
        aylik_min=deger2
        fark_min=sayac
    sayac=sayac+1

print(kur[fark_tarih][0],"/",kur[fark_tarih+fark][0],"Tarihleri Arasında",fark,"İş Günü Kadar Bekleyerek Yapabilecek En Yüksek Kâr: ",aylik_max)
print(kur[fark_min][0],"/",kur[fark_min+fark][0],"Tarihleri Arasında",fark,"İş Günü Kadar Bekleyerek Yapabilecek En Düşük Kâr: ",aylik_min)
